# -*- coding: utf-8 -*-

import base64
import csv
import io
from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

try:
    import openpyxl
    OPENPYXL_INSTALLED = True
except ImportError:
    OPENPYXL_INSTALLED = False


class FarmImportWizard(models.TransientModel):
    _name = 'farm.import.wizard'
    _description = 'معالج استيراد المزارع'

    file = fields.Binary(
        string='ملف Excel/CSV',
        required=True,
        help='اختر ملف Excel (xlsx) أو CSV يحتوي على بيانات المزارع',
    )
    filename = fields.Char(string='اسم الملف')
    
    import_log = fields.Text(
        string='سجل الاستيراد',
        readonly=True,
    )

    def _read_xlsx_file(self, file_data):
        """Read XLSX file and return list of dictionaries"""
        if not OPENPYXL_INSTALLED:
            raise UserError(_('مكتبة openpyxl غير مثبتة. الرجاء تثبيتها باستخدام: pip install openpyxl'))
        
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        
        headers = [str(h).strip() if h else '' for h in rows[0]]
        data = []
        
        for row in rows[1:]:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = str(value).strip() if value is not None else ''
            if any(row_dict.values()):  # Skip empty rows
                data.append(row_dict)
        
        return data

    def _read_csv_file(self, file_data):
        """Read CSV file and return list of dictionaries"""
        csv_file = io.StringIO(file_data.decode('utf-8-sig'))
        reader = csv.DictReader(csv_file)
        return list(reader)

    def action_import(self):
        """Import farms with hierarchy from Excel/CSV file"""
        self.ensure_one()
        
        if not self.file:
            raise UserError(_('الرجاء اختيار ملف'))
        
        # Decode file
        try:
            file_data = base64.b64decode(self.file)
            
            # Determine file type and read data
            if self.filename and self.filename.lower().endswith('.xlsx'):
                rows = self._read_xlsx_file(file_data)
            else:
                rows = self._read_csv_file(file_data)
        except Exception as e:
            raise UserError(_('خطأ في قراءة الملف: %s') % str(e))
        
        log_messages = []
        farms_created = 0
        sectors_created = 0
        units_created = 0
        houses_created = 0
        
        # Cache for already created records
        farm_cache = {}
        sector_cache = {}
        unit_cache = {}
        
        for row_num, row in enumerate(rows, start=2):
            try:
                # Get or create Farm
                farm_name = row.get('farm_name', '').strip()
                if not farm_name:
                    continue
                
                if farm_name not in farm_cache:
                    farm = self.env['farm.farm'].search([('name', '=', farm_name)], limit=1)
                    if not farm:
                        farm = self.env['farm.farm'].create({
                            'name': farm_name,
                            'code': row.get('farm_code', '').strip() or False,
                            'location': row.get('farm_location', '').strip() or False,
                        })
                        farms_created += 1
                        log_messages.append(f'✅ تم إنشاء المزرعة: {farm_name}')
                    farm_cache[farm_name] = farm.id
                
                farm_id = farm_cache[farm_name]
                
                # Get or create Sector
                sector_name = row.get('sector_name', '').strip()
                if sector_name:
                    sector_key = f"{farm_id}_{sector_name}"
                    if sector_key not in sector_cache:
                        sector = self.env['farm.sector'].search([
                            ('name', '=', sector_name),
                            ('farm_id', '=', farm_id)
                        ], limit=1)
                        if not sector:
                            sector = self.env['farm.sector'].create({
                                'name': sector_name,
                                'code': row.get('sector_code', '').strip() or False,
                                'farm_id': farm_id,
                            })
                            sectors_created += 1
                            log_messages.append(f'  ✅ تم إنشاء القطاع: {sector_name}')
                        sector_cache[sector_key] = sector.id
                    
                    sector_id = sector_cache[sector_key]
                    
                    # Get or create Unit
                    unit_name = row.get('unit_name', '').strip()
                    if unit_name:
                        unit_key = f"{sector_id}_{unit_name}"
                        if unit_key not in unit_cache:
                            unit = self.env['farm.unit'].search([
                                ('name', '=', unit_name),
                                ('sector_id', '=', sector_id)
                            ], limit=1)
                            if not unit:
                                unit = self.env['farm.unit'].create({
                                    'name': unit_name,
                                    'code': row.get('unit_code', '').strip() or False,
                                    'sector_id': sector_id,
                                })
                                units_created += 1
                                log_messages.append(f'    ✅ تم إنشاء الوحدة: {unit_name}')
                            unit_cache[unit_key] = unit.id
                        
                        unit_id = unit_cache[unit_key]
                        
                        # Get or create House
                        house_name = row.get('house_name', '').strip()
                        if house_name:
                            house = self.env['farm.house'].search([
                                ('name', '=', house_name),
                                ('unit_id', '=', unit_id)
                            ], limit=1)
                            if not house:
                                house_type = row.get('house_type', '').strip() or 'plastic'
                                if house_type not in ('glass', 'plastic', 'polycarbonate'):
                                    house_type = 'plastic'
                                
                                area = 0
                                try:
                                    area = float(row.get('house_area', '0').strip() or '0')
                                except:
                                    pass
                                
                                house = self.env['farm.house'].create({
                                    'name': house_name,
                                    'code': row.get('house_code', '').strip() or False,
                                    'unit_id': unit_id,
                                    'area': area,
                                    'house_type': house_type,
                                    'description': row.get('house_description', '').strip() or False,
                                })
                                houses_created += 1
                                log_messages.append(f'      ✅ تم إنشاء البيت: {house_name}')
                
            except Exception as e:
                log_messages.append(f'❌ خطأ في السطر {row_num}: {str(e)}')
        
        # Summary
        summary = f"""
╔══════════════════════════════════════╗
║         ملخص عملية الاستيراد          ║
╠══════════════════════════════════════╣
║  المزارع المنشأة:     {farms_created:>10}      ║
║  القطاعات المنشأة:    {sectors_created:>10}      ║
║  الوحدات المنشأة:     {units_created:>10}      ║
║  البيوت المنشأة:      {houses_created:>10}      ║
╚══════════════════════════════════════╝
"""
        log_messages.insert(0, summary)
        self.import_log = '\n'.join(log_messages)
        
        return {
            'type': 'ir.actions.act_window',
            'name': _('نتيجة الاستيراد'),
            'res_model': 'farm.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }


class ProjectImportWizard(models.TransientModel):
    _name = 'project.import.wizard'
    _description = 'معالج استيراد المشاريع'

    file = fields.Binary(
        string='ملف Excel/CSV',
        required=True,
        help='اختر ملف Excel (xlsx) أو CSV يحتوي على بيانات المشاريع',
    )
    filename = fields.Char(string='اسم الملف')
    
    import_log = fields.Text(
        string='سجل الاستيراد',
        readonly=True,
    )

    def _read_xlsx_file(self, file_data):
        """Read XLSX file and return list of dictionaries"""
        if not OPENPYXL_INSTALLED:
            raise UserError(_('مكتبة openpyxl غير مثبتة. الرجاء تثبيتها باستخدام: pip install openpyxl'))
        
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        
        headers = [str(h).strip() if h else '' for h in rows[0]]
        data = []
        
        for row in rows[1:]:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = str(value).strip() if value is not None else ''
            if any(row_dict.values()):  # Skip empty rows
                data.append(row_dict)
        
        return data

    def _read_csv_file(self, file_data):
        """Read CSV file and return list of dictionaries"""
        csv_file = io.StringIO(file_data.decode('utf-8-sig'))
        reader = csv.DictReader(csv_file)
        return list(reader)

    def action_import(self):
        """Import projects with house assignments from Excel/CSV file"""
        self.ensure_one()
        
        if not self.file:
            raise UserError(_('الرجاء اختيار ملف'))
        
        # Decode file
        try:
            file_data = base64.b64decode(self.file)
            
            # Determine file type and read data
            if self.filename and self.filename.lower().endswith('.xlsx'):
                rows = self._read_xlsx_file(file_data)
            else:
                rows = self._read_csv_file(file_data)
        except Exception as e:
            raise UserError(_('خطأ في قراءة الملف: %s') % str(e))
        
        log_messages = []
        projects_created = 0
        assignments_created = 0
        
        # Cache for already created records
        project_cache = {}
        
        for row_num, row in enumerate(rows, start=2):
            try:
                # Get Farm - search by code first, then by name
                farm_code = row.get('farm_code', '').strip()
                farm_name = row.get('farm_name', '').strip()
                
                if not farm_code and not farm_name:
                    log_messages.append(f'⚠️ السطر {row_num}: لم يتم تحديد المزرعة')
                    continue
                
                farm = None
                # Try to find farm by code first
                if farm_code:
                    farm = self.env['farm.farm'].search([('code', '=', farm_code)], limit=1)
                # Fall back to name if not found by code
                if not farm and farm_name:
                    farm = self.env['farm.farm'].search([('name', '=', farm_name)], limit=1)
                
                if not farm:
                    identifier = farm_code or farm_name
                    log_messages.append(f'❌ السطر {row_num}: المزرعة "{identifier}" غير موجودة')
                    continue
                
                # Get or create Project
                project_name = row.get('project_name', '').strip()
                if not project_name:
                    continue
                
                project_key = f"{farm.id}_{project_name}"
                if project_key not in project_cache:
                    project = self.env['farm.project'].search([
                        ('name', '=', project_name),
                        ('farm_id', '=', farm.id)
                    ], limit=1)
                    if not project:
                        project_vals = {
                            'name': project_name,
                            'farm_id': farm.id,
                            'notes': row.get('project_notes', '').strip() or False,
                        }
                        
                        # Parse dates
                        start_date = row.get('planned_start_date', '').strip()
                        if start_date:
                            try:
                                project_vals['planned_start_date'] = start_date
                            except:
                                pass
                        
                        end_date = row.get('expected_finish_date', '').strip()
                        if end_date:
                            try:
                                project_vals['expected_finish_date'] = end_date
                            except:
                                pass
                        
                        project = self.env['farm.project'].create(project_vals)
                        projects_created += 1
                        log_messages.append(f'✅ تم إنشاء المشروع: {project_name}')
                    project_cache[project_key] = project.id
                
                project_id = project_cache[project_key]
                
                # Get House - search by code first, then by name
                house_code = row.get('house_code', '').strip()
                house_name = row.get('house_name', '').strip()
                
                if house_code or house_name:
                    house = None
                    # Try to find house by code first
                    if house_code:
                        house = self.env['farm.house'].search([
                            ('code', '=', house_code),
                            ('farm_id', '=', farm.id)
                        ], limit=1)
                    # Fall back to name if not found by code
                    if not house and house_name:
                        house = self.env['farm.house'].search([
                            ('name', '=', house_name),
                            ('farm_id', '=', farm.id)
                        ], limit=1)
                    
                    if not house:
                        identifier = house_code or house_name
                        log_messages.append(f'  ⚠️ السطر {row_num}: البيت "{identifier}" غير موجود')
                        continue
                    
                    # Check if assignment exists
                    existing = self.env['farm.project.house'].search([
                        ('project_id', '=', project_id),
                        ('house_id', '=', house.id)
                    ], limit=1)
                    
                    if not existing:
                        assignment_vals = {
                            'project_id': project_id,
                            'house_id': house.id,
                            'season': row.get('season', '').strip() or False,
                            'activity_description': row.get('activity_description', '').strip() or False,
                        }
                        
                        # Get Product - search by code first, then by name
                        product_code = row.get('product_code', '').strip()
                        product_name = row.get('product_name', '').strip()
                        
                        if product_code or product_name:
                            product = None
                            # Try to find product by code (default_code) first
                            if product_code:
                                product = self.env['product.product'].search([
                                    ('default_code', '=', product_code)
                                ], limit=1)
                            # Fall back to name if not found by code
                            if not product and product_name:
                                product = self.env['product.product'].search([
                                    ('name', '=', product_name)
                                ], limit=1)
                            if product:
                                assignment_vals['product_id'] = product.id
                        
                        # Expected quantity
                        try:
                            expected_qty = float(row.get('expected_qty', '0').strip() or '0')
                            assignment_vals['expected_qty'] = expected_qty
                        except:
                            pass
                        
                        # UoM
                        uom_name = row.get('uom_name', '').strip()
                        if uom_name:
                            uom = self.env['uom.uom'].search([
                                ('name', '=', uom_name)
                            ], limit=1)
                            if uom:
                                assignment_vals['uom_id'] = uom.id
                        
                        self.env['farm.project.house'].create(assignment_vals)
                        assignments_created += 1
                        house_identifier = house_code or house_name
                        log_messages.append(f'  ✅ تم تخصيص البيت: {house_identifier}')
                
            except Exception as e:
                log_messages.append(f'❌ خطأ في السطر {row_num}: {str(e)}')
        
        # Summary
        summary = f"""
╔══════════════════════════════════════╗
║         ملخص عملية الاستيراد          ║
╠══════════════════════════════════════╣
║  المشاريع المنشأة:    {projects_created:>10}      ║
║  تخصيصات البيوت:     {assignments_created:>10}      ║
╚══════════════════════════════════════╝
"""
        log_messages.insert(0, summary)
        self.import_log = '\n'.join(log_messages)
        
        return {
            'type': 'ir.actions.act_window',
            'name': _('نتيجة الاستيراد'),
            'res_model': 'project.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

