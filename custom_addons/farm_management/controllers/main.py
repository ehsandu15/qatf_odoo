# -*- coding: utf-8 -*-

import io
import os
from odoo import http
from odoo.http import request, content_disposition

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_INSTALLED = True
except ImportError:
    OPENPYXL_INSTALLED = False


class FarmTemplateController(http.Controller):
    
    def _create_xlsx_template(self, headers, sample_data, sheet_name='Data'):
        """Create XLSX template with headers and sample data"""
        if not OPENPYXL_INSTALLED:
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        # Write sample data
        for row_num, row_data in enumerate(sample_data, start=2):
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _get_farm_template_data(self):
        """Farm complete import template data"""
        headers = [
            'farm_name', 'farm_code', 'farm_location',
            'sector_name', 'sector_code',
            'unit_name', 'unit_code',
            'house_name', 'house_code', 'house_area', 'house_type', 'house_description'
        ]
        sample_data = [
            ['مزرعة النخيل', 'F001', 'الرياض', 'القطاع الشمالي', 'S001', 'الوحدة 1', 'U001', 'بيت 1', 'H001', '500', 'glass', 'بيت زجاجي'],
            ['مزرعة النخيل', 'F001', 'الرياض', 'القطاع الشمالي', 'S001', 'الوحدة 1', 'U001', 'بيت 2', 'H002', '450', 'plastic', 'بيت بلاستيكي'],
            ['مزرعة النخيل', 'F001', 'الرياض', 'القطاع الشمالي', 'S001', 'الوحدة 2', 'U002', 'بيت 3', 'H003', '600', 'polycarbonate', 'بيت بوليكربونيت'],
            ['مزرعة النخيل', 'F001', 'الرياض', 'القطاع الجنوبي', 'S002', 'الوحدة 3', 'U003', 'بيت 4', 'H004', '550', 'glass', ''],
        ]
        return headers, sample_data

    def _get_project_template_data(self):
        """Project complete import template data"""
        headers = [
            'project_name', 'farm_name', 'farm_code', 'planned_start_date', 'expected_finish_date', 'project_notes',
            'house_name', 'house_code', 'product_name', 'product_code', 'expected_qty', 'uom_name', 'season', 'activity_description'
        ]
        sample_data = [
            ['مشروع طماطم 2024', 'مزرعة النخيل', 'F001', '2024-01-01', '2024-06-30', 'مشروع إنتاج الطماطم', 'بيت 1', 'H001', 'طماطم', '7001', '5000', 'كغ', 'الربيع', 'زراعة طماطم'],
            ['مشروع طماطم 2024', 'مزرعة النخيل', 'F001', '2024-01-01', '2024-06-30', '', 'بيت 2', 'H002', 'طماطم', '7001', '4500', 'كغ', 'الربيع', ''],
            ['مشروع خيار 2024', 'مزرعة النخيل', 'F001', '2024-02-01', '2024-05-31', 'مشروع إنتاج الخيار', 'بيت 3', 'H003', 'خيار', '7002', '3000', 'كغ', 'الصيف', 'زراعة خيار'],
        ]
        return headers, sample_data

    @http.route('/farm_management/download_template/<string:template_name>', type='http', auth='user')
    def download_template(self, template_name, **kwargs):
        """Download import template as XLSX file"""
        
        # Template configurations
        templates = {
            'farm_complete_import.xlsx': ('farm', 'قالب_المزارع'),
            'project_complete_import.xlsx': ('project', 'قالب_المشاريع'),
            # Legacy CSV names - redirect to XLSX
            'farm_complete_import.csv': ('farm', 'قالب_المزارع'),
            'project_complete_import.csv': ('project', 'قالب_المشاريع'),
        }
        
        if template_name not in templates:
            return request.not_found()
        
        if not OPENPYXL_INSTALLED:
            # Fallback to CSV if openpyxl not installed
            return self._download_csv_template(template_name)
        
        template_type, sheet_name = templates[template_name]
        
        if template_type == 'farm':
            headers, sample_data = self._get_farm_template_data()
        else:
            headers, sample_data = self._get_project_template_data()
        
        file_content = self._create_xlsx_template(headers, sample_data, sheet_name)
        
        if not file_content:
            return request.not_found()
        
        # Always return as XLSX
        xlsx_filename = template_name.replace('.csv', '.xlsx')
        if not xlsx_filename.endswith('.xlsx'):
            xlsx_filename += '.xlsx'
        
        return request.make_response(
            file_content,
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition(xlsx_filename)),
            ]
        )

    def _download_csv_template(self, template_name):
        """Fallback to download CSV template if openpyxl not installed"""
        # Get module path
        module_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        file_path = os.path.join(module_path, 'static', 'import_templates', template_name)
        
        if not os.path.exists(file_path):
            return request.not_found()
        
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        return request.make_response(
            file_content,
            headers=[
                ('Content-Type', 'text/csv; charset=utf-8'),
                ('Content-Disposition', content_disposition(template_name)),
            ]
        )

